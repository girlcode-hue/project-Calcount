()
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
import matplotlib.pyplot as plt

user_database = Workbook()

def new_sheet(username):
    user_database.create_sheet(username)

nut= load_workbook(filename='C:/Users/91960/Downloads/nutrition.xlsx',read_only=False)
North = nut['North']
South = nut['South']
user_database = load_workbook('C:/Users/91960/Downloads/User_data_file.xlsx')

def new_user(user_database):
    user_name = input("enter username: ")
    new_sheet(username=user_name)
    user_database[user_name]['A1'] = 'Day'
    user_database[user_name]['B1'] = 'Breakfast'
    user_database[user_name]['C1'] = 'Lunch'
    user_database[user_name]['D1'] = 'Dinner'
    user_database[user_name]['E1'] = 'Total Calories'
    user_database[user_name]['F1'] = 'Average Calories'

    user_database.save('C:/Users/91960/Downloads/User_data_file.xlsx')
    print(("hello {} !").format(user_name))

#new_user(user_database)

def tot_cal_mark1(num_food_items, cuisine):
    for food_num in range(num_food_items):
      globals()["food_" + str(food_num)] = [input("Enter the food item name: "), int(input("Number of servings: "))]
    start = 0
    for i in range(2,len(cuisine['A'])):
      for j in range(num_food_items):
        if globals()["food_" + str(j)][0] == cuisine['A'+str(i+1)].value:
            start += globals()["food_" + str(j)][1]*cuisine['B'+str(i+1)].value
    print('the number of calories consumed is, {}'.format(start) + ' KCal')
    return start


def cal_comp(cuisine,username):
    day = int(input("enter day number: "))
    print("good morning. Hope you had a good breakfast")
    num_food = int(input("how many food items did you eat? "))
    breakfast = tot_cal_mark1(num_food_items=num_food,cuisine=cuisine)
    user_database[username]['A'+str(day+1)].value = day
    user_database[username]['B'+str(day+1)].value = breakfast
    print("ok! tell us what you had for lunch!")
    num_food = int(input("how many food items did you eat? "))
    lunch = tot_cal_mark1(num_food_items=num_food,cuisine=cuisine)
    print("cool.and for dinner?")
    num_food = int(input("how many food items did you eat? "))
    dinner = tot_cal_mark1(num_food_items=num_food,cuisine=cuisine)
    user_database[username]['C'+str(day+1)].value = lunch
    user_database[username]['D'+str(day+1)].value = dinner
    tot_cal = breakfast+lunch+dinner
    print('total calorie for the day is {}'.format(tot_cal))
    avrg_cal = tot_cal/3
    print('average calories consumed is {}'.format(avrg_cal))
    user_database[username]['E'+str(day+1)].value = tot_cal
    user_database[username]['F'+str(day+1)].value = avrg_cal

    user_database.save('C:/Users/91960/Downloads/User_data_file.xlsx')

#print(cal_comp('South','darsna19'))

def cal_plot():
    names = []
    tot_cal =[]
    for i in range(2,len(user_database['darsna19']['A'])):
        names.append(i)

    for j in range(2,len(user_database['darsna19']['E'])):
        tot_cal.append(user_database['darsna19']['E'+str(j+1)].value)

    print(names,tot_cal)

    plt.plot(names,tot_cal)
    plt.show()

cal_plot()